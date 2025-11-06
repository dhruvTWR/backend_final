"""
Complete Flask Server - Face Recognition Attendance System
Uses all your existing models
"""

from flask import Flask, request, jsonify, send_file, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime,date, timedelta


# Import models (tumhare jo files hain)
from models.student import Student
from models.teacher import Teacher
from models.admin import Admin
from models.subjects import Subject
from models.class_subject import ClassSubject
from models.attendance import Attendance
from models.unrecognized_face import UnrecognizedFace 
from models.log import Log

# Import services
from services.face_recognition_service import FaceRecognitionService
from utils.logger import setup_logger
from utils.excel_export import ExcelExporter
from config import Config
from database.db import Database 
from utils.image_quality_checker import ImageQualityChecker


# Initialize Flask
app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
CORS(
    app,
    supports_credentials=True,
    origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://192.168.80.157:3000/",
        "http://localhost:5173",
        "http://192.168.1.9:5173",
        "http://192.168.80.157:5000/",
        "http://192.168.80.157:3000/",
        "http://localhost:3001/teacher/dashboard/"
        
    ]
)

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize models
student_model = Student()
teacher_model = Teacher()
admin_model = Admin()
subject_model = Subject()
class_subject_model = ClassSubject()
attendance_model = Attendance()
face_service = FaceRecognitionService()
unrecognized_model = UnrecognizedFace()
log_model = Log()

quality_checker = ImageQualityChecker()



# Create directories
os.makedirs('uploads/attendance_images', exist_ok=True)
os.makedirs('uploads/student_photos', exist_ok=True)
os.makedirs('uploads/unrecognized_faces', exist_ok=True)
os.makedirs('exports', exist_ok=True)


# ============================================================================
# AUTHENTICATION
# ============================================================================

@app.route('/api/auth/admin/login', methods=['POST'])
def admin_login():
    """Admin login"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        admin = admin_model.authenticate(username, password)
        
        if admin:
            session['user_id'] = admin['id']
            session['user_type'] = 'admin'
            session['username'] = admin['username']
            
            return jsonify({'success': True, 'message': 'Login successful', 'user': admin}), 200
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
    except Exception as e:
        logger.error(f"Admin login error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/auth/teacher/login', methods=['POST'])
def teacher_login():
    """Teacher login"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        teacher = teacher_model.authenticate(username, password)
        
        if teacher:
            session['user_id'] = teacher['id']
            session['user_type'] = 'teacher'
            session['username'] = teacher['username']
            
            return jsonify({'success': True, 'message': 'Login successful', 'user': teacher}), 200
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
    except Exception as e:
        logger.error(f"Teacher login error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Logout"""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out'}), 200


@app.route('/api/auth/verify', methods=['GET'])
def verify_session():
    """Verify session"""
    if 'user_id' in session:
        return jsonify({
            'success': True,
            'authenticated': True,
            'user': {
                'id': session['user_id'],
                'username': session.get('username'),
                'type': session.get('user_type')
            }
        }), 200
    return jsonify({'success': True, 'authenticated': False}), 200


# ============================================================================
# ADMIN - STUDENTS
# ============================================================================

@app.route('/api/admin/students', methods=['GET'])
def admin_get_students():
    """Get students"""
    try:
        branch_id = request.args.get('branch_id', type=int)
        year = request.args.get('year', type=int)
        section = request.args.get('section')
        
        if branch_id and year and section:
            students = student_model.get_students_by_class(branch_id, year, section)
        else:
            students = student_model.get_all_students(branch_id, year)
        
        return jsonify({'success': True, 'count': len(students), 'students': students}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/students', methods=['POST'])
def admin_add_student():
    """Add student with photo"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'Photo required'}), 400
        
        file = request.files['file']
        roll_number = request.form.get('roll_number')
        name = request.form.get('name')
        email = request.form.get('email')
        branch_id = request.form.get('branch_id', type=int)
        year = request.form.get('year', type=int)
        section = request.form.get('section')
        
        if not all([roll_number, name, branch_id, year, section]):
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Save photo
        filename = secure_filename(f"{roll_number}_{file.filename}")
        photo_path = os.path.join('uploads/student_photos', filename)
        file.save(photo_path)
        
        # Generate encoding
        encoding = face_service.generate_encoding(photo_path)
        if encoding is None:
            os.remove(photo_path)
            return jsonify({'success': False, 'message': 'No face detected'}), 400
        
        # Add student
        student_id = student_model.add_student(
            roll_number=roll_number,
            name=name,
            email=email,
            branch_id=branch_id,
            year=year,
            section=section,
            face_encoding=encoding,
            photo_path=photo_path
        )
        
        return jsonify({
            'success': True,
            'message': 'Student added successfully',
            'student_id': student_id
        }), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/students/<int:student_id>', methods=['PUT'])
def admin_update_student(student_id):
    """Update student"""
    try:
        data = request.form
        update_data = {}
        
        if data.get('name'):
            update_data['name'] = data.get('name')
        if data.get('email'):
            update_data['email'] = data.get('email')
        
        if 'file' in request.files:
            file = request.files['file']
            filename = secure_filename(file.filename)
            photo_path = os.path.join('uploads/student_photos', filename)
            file.save(photo_path)
            
            encoding = face_service.generate_encoding(photo_path)
            if encoding:
                update_data['photo_path'] = photo_path
                update_data['face_encoding'] = encoding
        
        student_model.update_student(student_id, **update_data)
        return jsonify({'success': True, 'message': 'Student updated'}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/students/<int:student_id>', methods=['DELETE'])
def admin_delete_student(student_id):
    """Delete student"""
    try:
        student_model.delete_student(student_id, soft_delete=True)
        return jsonify({'success': True, 'message': 'Student deleted'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/students/search', methods=['GET'])
def admin_search_students():
    """Search students"""
    try:
        query = request.args.get('q', '')
        students = student_model.search_students(query)
        return jsonify({'success': True, 'students': students}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# ADMIN - TEACHERS
# ============================================================================

@app.route('/api/admin/teachers', methods=['GET'])
def admin_get_teachers():
    """Get all teachers"""
    try:
        teachers = teacher_model.get_all_teachers()
        return jsonify({'success': True, 'teachers': teachers}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/teachers', methods=['POST'])
def admin_add_teacher():
    """Add teacher"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        teacher_id = teacher_model.add_teacher(username, password)
        return jsonify({'success': True, 'message': 'Teacher added', 'teacher_id': teacher_id}), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/teachers/<int:teacher_id>', methods=['PUT'])
def admin_update_teacher(teacher_id):
    """Update teacher"""
    try:
        data = request.get_json()
        teacher_model.update_teacher(teacher_id, **data)
        return jsonify({'success': True, 'message': 'Teacher updated'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# ADMIN - SUBJECTS
# ============================================================================

@app.route('/api/admin/subjects', methods=['GET'])
def admin_get_subjects():
    """Get all subjects"""
    try:
        subjects = subject_model.get_all_subjects()
        return jsonify({'success': True, 'subjects': subjects}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/subjects', methods=['POST'])
def admin_add_subject():
    """Add subject"""
    try:
        data = request.get_json()
        subject_code = data.get('subject_code')
        subject_name = data.get('subject_name')
        
        if not subject_code or not subject_name:
            return jsonify({'success': False, 'message': 'Code and name required'}), 400
        
        subject_id = subject_model.add_subject(subject_code, subject_name)
        return jsonify({'success': True, 'message': 'Subject added', 'subject_id': subject_id}), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/subjects/<int:subject_id>', methods=['DELETE'])
def admin_delete_subject(subject_id):
    """Delete subject"""
    try:
        subject_model.delete_subject(subject_id)
        return jsonify({'success': True, 'message': 'Subject deleted'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/subjects/search', methods=['GET'])
def admin_search_subjects():
    """Search subjects"""
    try:
        query = request.args.get('q', '')
        subjects = subject_model.search_subjects(query)
        return jsonify({'success': True, 'subjects': subjects}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# ADMIN - CLASS SUBJECTS
# ============================================================================



@app.route('/api/admin/class-subjects', methods=['GET'])
def admin_get_class_subjects():
    """Get all class-subject mappings"""
    try:
        academic_year = request.args.get('academic_year')
        branch_id = request.args.get('branch_id', type=int)
        
        mappings = class_subject_model.get_all_mappings(academic_year, branch_id)
        
        # Transform the response to match React component expectations
        formatted_mappings = []
        for mapping in mappings:
            formatted_mappings.append({
                'class_subject_id': mapping['id'],
                'subject_id': mapping.get('subject_id'),  # Add this if available
                'branch_id': mapping['branch_id'],
                'year': mapping['year'],
                'section': mapping['section'],
                'semester': mapping['semester'],
                'academic_year': mapping['academic_year'],
                'is_active': mapping['is_active'],
                'subject_name': mapping['subject_name'],
                'subject_code': mapping['subject_code'],
                'branch_name': mapping['branch_name'],
                'branch_code': mapping['branch_code'],
                'created_at': mapping['created_at']
            })
        
        return jsonify({'success': True, 'mappings': formatted_mappings}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/class-subjects', methods=['POST'])
def admin_add_class_subject():
    """Assign subject to class"""
    try:
        data = request.get_json()
        
        required = ['subject_id', 'branch_id', 'year', 'section', 'semester', 'academic_year']
        if not all(data.get(f) for f in required):
            return jsonify({'success': False, 'message': 'Missing fields'}), 400
        
        class_subject_id = class_subject_model.assign_subject_to_class(
            subject_id=data['subject_id'],
            branch_id=data['branch_id'],
            year=data['year'],
            section=data['section'],
            semester=data['semester'],
            academic_year=data['academic_year']
        )
        
        return jsonify({
            'success': True,
            'message': 'Subject assigned to class',
            'class_subject_id': class_subject_id
        }), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/class-subjects/<int:class_subject_id>', methods=['PUT'])
def admin_update_class_subject(class_subject_id):
    """Update class-subject mapping"""
    try:
        data = request.get_json()
        class_subject_model.update_class_subject(
            class_subject_id,
            semester=data.get('semester'),
            academic_year=data.get('academic_year')
        )
        return jsonify({'success': True, 'message': 'Mapping updated'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/admin/class-subjects/<int:class_subject_id>', methods=['DELETE'])
def admin_delete_class_subject(class_subject_id):
    """Delete class-subject mapping"""
    try:
        class_subject_model.delete_class_subject(class_subject_id)
        return jsonify({'success': True, 'message': 'Mapping deleted'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# TEACHER - GET SUBJECTS FOR CLASS
# ============================================================================

@app.route('/api/teacher/class-subjects', methods=['GET'])
def teacher_get_subjects():
    """
    Get subjects for selected class
    Query: ?branch_id=1&year=3&section=A
    """
    try:
        branch_id = request.args.get('branch_id', type=int)
        year = request.args.get('year', type=int)
        section = request.args.get('section')
        academic_year = request.args.get('academic_year')
        
        if not all([branch_id, year, section]):
            return jsonify({'success': False, 'message': 'branch_id, year, section required'}), 400
        
        subjects = class_subject_model.get_subjects_for_class(
            branch_id, year, section, academic_year
        )
        
        return jsonify({'success': True, 'subjects': subjects}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# IMAGE QUALITY CHECK ENDPOINT 
# ============================================================================

@app.route('/api/teacher/attendance/check-quality', methods=['POST'])
def check_image_quality():
    """
    Check image quality before processing attendance
    Allows teacher to verify image is acceptable
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Save temporary file
        filename = secure_filename(f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        temp_path = os.path.join('uploads/attendance_images', filename)
        file.save(temp_path)
        
        # Check quality
        quality_result = face_service.check_image_quality(temp_path)
        
        # Get face count for preview
        # face_count = face_service.get_face_count(temp_path)
        # quality_result['face_count'] = face_count
        
        # Keep temp file for potential use
        quality_result['temp_path'] = filename
        
        return jsonify({
            'success': True,
            'quality_check': quality_result,
            'temp_filename': filename
        }), 200
        
    except Exception as e:
        logger.error(f"Quality check error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# BATCH/MULTIPLE IMAGE ATTENDANCE MARKING
# ============================================================================

@app.route('/api/teacher/attendance/mark-batch', methods=['POST'])
def teacher_mark_attendance_batch():
    """
    Mark attendance using multiple images
    Allows teachers to upload photos from different angles/rows
    """
    try:
        # Check if files exist
        if 'files' not in request.files:
            return jsonify({'success': False, 'message': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        
        if not files or len(files) == 0:
            return jsonify({'success': False, 'message': 'No files selected'}), 400
        
        # Get parameters
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        skip_quality_check = request.form.get('skip_quality_check', 'false').lower() == 'true'
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        logger.info(f"Processing batch attendance for class_subject_id: {class_subject_id}")
        logger.info(f"Number of images: {len(files)}")
        
        # Parse date
        if isinstance(attendance_date_str, str):
            attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
        else:
            attendance_date = attendance_date_str
        
        # Get class info
        class_info = class_subject_model.get_class_subject_by_id(class_subject_id)
        
        if not class_info:
            return jsonify({'success': False, 'message': 'Class subject not found'}), 404
        
        # Save all uploaded files
        saved_files = []
        quality_checks = []
        
        for idx, file in enumerate(files):
            filename = secure_filename(f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{idx}_{file.filename}")
            filepath = os.path.join('uploads/attendance_images', filename)
            file.save(filepath)
            
            # Validate image
            import cv2
            test_img = cv2.imread(filepath)
            if test_img is None:
                os.remove(filepath)
                quality_checks.append({
                    'filename': file.filename,
                    'acceptable': False,
                    'error': 'Invalid image format'
                })
                continue
            
            # Check quality (unless skipped)
            if not skip_quality_check:
                quality_result = face_service.check_image_quality(filepath)
                quality_checks.append({
                    'filename': file.filename,
                    **quality_result
                })
                
                if not quality_result['acceptable']:
                    logger.warning(f"Image {file.filename} failed quality check")
                    # Don't delete yet - let teacher decide
                    continue
            
            saved_files.append(filepath)
            logger.info(f"Saved file {idx+1}/{len(files)}: {filepath}")
        
        if not saved_files:
            return jsonify({
                'success': False,
                'message': 'No acceptable images found',
                'quality_checks': quality_checks
            }), 400
        
        # Process all acceptable images together
        result = face_service.batch_recognize_students(
            image_paths=saved_files,
            branch_id=class_info['branch_id'],
            year=class_info['year'],
            section=class_info['section'],
            class_subject_id=class_subject_id
        )
        
        logger.info(f"Batch recognition result: {result}")
        
        # Mark attendance for all recognized students
        attendance_ids = []
        attendance_time = datetime.now().time()
        
        for student in result.get('recognized', []):
            try:
                attendance_id = attendance_model.mark_attendance(
                    student_id=student['student_id'],
                    class_subject_id=class_subject_id,
                    attendance_date=attendance_date,
                    attendance_time=attendance_time,
                    status='present',
                    confidence=student.get('confidence', 0.0),
                    image_path=saved_files[0],  # Reference first image
                    marked_by=session.get('user_id'),
                    remarks=f'Batch recognition ({len(saved_files)} images)'
                )
                attendance_ids.append(attendance_id)
                logger.info(f"Marked attendance for {student['name']}: {attendance_id}")
            except Exception as e:
                logger.error(f"Error marking attendance for {student.get('name')}: {e}")
        
        # Save unrecognized faces
        unrecognized_count = 0
        for unrec in result.get('unrecognized', []):
            try:
                unrec_id = unrecognized_model.add_unrecognized(
                    class_subject_id=class_subject_id,
                    status='absent',
                    image_path=unrec.get('image_path', ''),
                    original_upload=unrec.get('source_image', ''),
                    attendance_date=attendance_date
                )
                if unrec_id:
                    unrecognized_count += 1
            except Exception as e:
                logger.error(f"Failed to save unrecognized face: {e}")
        
        return jsonify({
            'success': True,
            'message': f'Batch attendance marked. {len(attendance_ids)} students present.',
            'images_processed': result.get('images_processed', 0),
            'images_failed': result.get('images_failed', 0),
            'total_faces': result.get('total_faces', 0),
            'recognized_count': len(result.get('recognized', [])),
            'unrecognized_count': len(result.get('unrecognized', [])),
            'students': result.get('recognized', []),
            'attendance_ids': attendance_ids,
            'quality_checks': quality_checks,
            'failed_images': result.get('failed_images', [])
        }), 200
        
    except Exception as e:
        logger.error(f"Batch attendance error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500


# ============================================================================
# ENHANCED SINGLE IMAGE ATTENDANCE WITH QUALITY CHECK
# ============================================================================

@app.route('/api/teacher/attendance/mark-v2', methods=['POST'])
def teacher_mark_attendance_v2():
    """
    Enhanced version with automatic quality check
    Replaces old /mark endpoint
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        force_process = request.form.get('force_process', 'false').lower() == 'true'
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        # Parse date
        if isinstance(attendance_date_str, str):
            attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
        else:
            attendance_date = attendance_date_str
        
        # Get class info
        class_info = class_subject_model.get_class_subject_by_id(class_subject_id)
        
        if not class_info:
            return jsonify({'success': False, 'message': 'Class subject not found'}), 404
        
        # Save file
        filename = secure_filename(f"att_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join('uploads/attendance_images', filename)
        file.save(filepath)
        
        # Validate image
        import cv2
        test_img = cv2.imread(filepath)
        if test_img is None:
            os.remove(filepath)
            return jsonify({'success': False, 'message': 'Invalid image file'}), 400
        
        # Check quality first
        quality_check = face_service.check_image_quality(filepath)
        
        if not quality_check['acceptable'] and not force_process:
            return jsonify({
                'success': False,
                'message': 'Image quality not acceptable',
                'quality_check': quality_check,
                'can_force': True,  # Allow teacher to override
                'temp_filename': filename
            }), 400
        
        # Process attendance
        result = face_service.recognize_students(
            image_path=filepath,
            branch_id=class_info['branch_id'],
            year=class_info['year'],
            section=class_info['section'],
            class_subject_id=class_subject_id
        )
        
        # Mark attendance
        attendance_ids = []
        attendance_time = datetime.now().time()
        
        for student in result.get('recognized', []):
            try:
                attendance_id = attendance_model.mark_attendance(
                    student_id=student['student_id'],
                    class_subject_id=class_subject_id,
                    attendance_date=attendance_date,
                    attendance_time=attendance_time,
                    status='present',
                    confidence=student.get('confidence', 0.0),
                    image_path=filepath,
                    marked_by=session.get('user_id'),
                    remarks='Face recognition'
                )
                attendance_ids.append(attendance_id)
            except Exception as e:
                logger.error(f"Error marking attendance: {e}")
        
        # Save unrecognized
        unrecognized_count = 0
        for unrec in result.get('unrecognized', []):
            try:
                unrec_id = unrecognized_model.add_unrecognized(
                    class_subject_id=class_subject_id,
                    status='absent',
                    image_path=unrec.get('image_path', ''),
                    original_upload=filepath,
                    attendance_date=attendance_date
                )
                if unrec_id:
                    unrecognized_count += 1
            except Exception as e:
                logger.error(f"Failed to save unrecognized: {e}")
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked. {len(attendance_ids)} students present.',
            'total_faces': result.get('total_faces', 0),
            'recognized_count': len(result.get('recognized', [])),
            'unrecognized_count': len(result.get('unrecognized', [])),
            'students': result.get('recognized', []),
            'attendance_ids': attendance_ids,
            'quality_check': quality_check,
            'was_forced': force_process
        }), 200
        
    except Exception as e:
        logger.error(f"Mark attendance v2 error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# TEACHER - MARK ATTENDANCE
# ============================================================================

@app.route('/api/teacher/attendance/mark', methods=['POST'])
def teacher_mark_attendance():
    
    """
    Mark attendance by uploading photo
    Form data: file, class_subject_id, attendance_date
    """
    try:
        # Validate file
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Get parameters
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        logger.info(f"Processing attendance for class_subject_id: {class_subject_id}")
        
        # Parse date
        if isinstance(attendance_date_str, str):
            attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
        else:
            attendance_date = attendance_date_str
        
        class_info = class_subject_model.get_class_subject_by_id(class_subject_id)
        
        if not class_info:
            return jsonify({'success': False, 'message': f'Class subject ID {class_subject_id} not found'}), 404
        
        logger.info(f"Class info: {class_info}")
        
        # Verify required fields exist
        required_fields = ['branch_id', 'year', 'section']
        missing = [f for f in required_fields if f not in class_info]
        if missing:
            return jsonify({
                'success': False, 
                'message': f'Missing fields in class_info: {missing}'
            }), 500
        
        # Save uploaded file
        filename = secure_filename(f"att_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join('uploads/attendance_images', filename)
        file.save(filepath)
        logger.info(f"File saved: {filepath}")
        
        # Validate image can be read
        import cv2
        test_img = cv2.imread(filepath)
        if test_img is None:
            os.remove(filepath)
            return jsonify({'success': False, 'message': 'Invalid image file'}), 400
        
        logger.info(f"Image size: {test_img.shape}")
        
        result = face_service.recognize_students(
            image_path=filepath,
            branch_id=class_info['branch_id'],
            year=class_info['year'],
            section=class_info['section'],
            class_subject_id=class_subject_id
        )
        
        logger.info(f"Recognition result: {result}")
        
        # Check if error occurred
        if 'error' in result:
            return jsonify({
                'success': False,
                'message': result['error'],
                'details': result
            }), 500
        
        attendance_ids = []
        attendance_time = datetime.now().time()
        
        for student in result.get('recognized', []):
            try:
                attendance_id = attendance_model.mark_attendance(
                    student_id=student['student_id'],
                    class_subject_id=class_subject_id,
                    attendance_date=attendance_date,
                    attendance_time=attendance_time,
                    status='present',
                    confidence=student.get('confidence', 0.0),
                    image_path=filepath,
                    marked_by=session.get('user_id'),
                    remarks='Face recognition'
                )
                attendance_ids.append(attendance_id)
                logger.info(f"Marked attendance for {student['name']}: {attendance_id}")
            except Exception as e:
                logger.error(f"Error marking attendance for student {student.get('name')}: {e}")
        
        unrecognized_count = 0
        for unrec in result.get('unrecognized', []):
            try:
                from models.unrecognized_face import UnrecognizedFace
                unrecognized_model = UnrecognizedFace()
        
                unrec_id = unrecognized_model.add_unrecognized(
                class_subject_id=class_subject_id,
                status='absent',
                image_path=unrec.get('image_path', ''),
                original_upload=filepath,
                attendance_date=attendance_date
                )
        
                if unrec_id:
                    unrecognized_count += 1
                    logger.info(f"Saved unrecognized face ID: {unrec_id}")
        
            except Exception as e:
                logger.error(f"Failed to save unrecognized face: {e}")
                import traceback
                traceback.print_exc()

        logger.info(f"Saved {unrecognized_count} unrecognized faces to database")
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked successfully. {len(attendance_ids)} students present.',
            'total_faces': result.get('total_faces', 0),
            'recognized_count': len(result.get('recognized', [])),
            'unrecognized_count': len(result.get('unrecognized', [])),
            'students': result.get('recognized', []),
            'attendance_ids': attendance_ids,
            'class_info': {
                'branch_id': class_info['branch_id'],
                'year': class_info['year'],
                'section': class_info['section']
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Mark attendance error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False, 
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500




@app.route('/api/teacher/attendance/manual', methods=['POST'])
def teacher_mark_manual():
    """Manually mark attendance for one student"""
    try:
        data = request.get_json()
        
        attendance_id = attendance_model.mark_attendance(
            student_id=data['student_id'],
            class_subject_id=data['class_subject_id'],
            attendance_date=datetime.strptime(data.get('attendance_date', date.today().isoformat()), '%Y-%m-%d').date(),
            attendance_time=datetime.now().time(),
            status=data.get('status', 'present'),
            marked_by=session.get('user_id'),
            remarks=data.get('remarks', 'Manually marked')
        )
        
        return jsonify({'success': True, 'attendance_id': attendance_id}), 201
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# TEACHER - ATTENDANCE REPORTS
# ============================================================================

@app.route('/api/teacher/attendance/report', methods=['GET'])
def teacher_get_report():
    """Get attendance report"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        report = attendance_model.get_attendance_by_date_range(
            class_subject_id, start_date, end_date
        )
        
        return jsonify({'success': True, 'report': report}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/teacher/attendance/by-date', methods=['GET'])
def teacher_get_by_date():
    """Get attendance for specific date"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        attendance_date = request.args.get('date', date.today().isoformat())
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        records = attendance_model.get_attendance_by_date(class_subject_id, attendance_date)
        return jsonify({'success': True, 'records': records}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# DROPDOWNS & UTILITY
# ============================================================================

@app.route('/api/branches', methods=['GET'])
def get_branches():
    """Get branches"""
    try:
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM branches ORDER BY branch_name")
        branches = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'branches': branches}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/dropdowns/years', methods=['GET'])
def get_years():
    return jsonify({'success': True, 'years': [1, 2, 3, 4]}), 200


@app.route('/api/dropdowns/sections', methods=['GET'])
def get_sections():
    return jsonify({'success': True, 'sections': ['A', 'B']}), 200


@app.route('/api/dropdowns/semesters', methods=['GET'])
def get_semesters():
    return jsonify({'success': True, 'semesters': list(range(1, 9))}), 200


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()}), 200


@app.route('/')
def index():
    return jsonify({
        'message': 'Face Recognition Attendance System API',
        'version': '1.0',
        'status': 'running'
    }), 200







## adons logs and unrecognize 


# ============================================================================
# UNRECOGNIZED FACES MANAGEMENT
# ============================================================================

@app.route('/api/unrecognized-faces', methods=['GET'])
def get_unrecognized_faces():
    """Get unrecognized faces with filters"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        resolved = request.args.get('resolved', 'false').lower() == 'true'
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 100, type=int)
        
        from models.unrecognized_face import UnrecognizedFace
       
        
        faces = unrecognized_model.get_unrecognized_faces(
            class_subject_id=class_subject_id,
            resolved=resolved,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )
        
        return jsonify({'success': True, 'count': len(faces), 'faces': faces}), 200
    except Exception as e:
        logger.error(f"Error getting unrecognized faces: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/count', methods=['GET'])
def get_unresolved_count():
    """Get count of unresolved faces"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        count = unrecognized_model.get_unresolved_count(class_subject_id)
        return jsonify({'success': True, 'count': count}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/recent', methods=['GET'])
def get_recent_unresolved():
    """Get recent unresolved faces"""
    try:
        days = request.args.get('days', 7, type=int)
        limit = request.args.get('limit', 50, type=int)
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        faces = unrecognized_model.get_recent_unresolved(days=days, limit=limit)
        return jsonify({'success': True, 'count': len(faces), 'faces': faces}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/by-date', methods=['GET'])
def get_unrecognized_by_date():
    """Get unrecognized faces for specific date"""
    try:
        attendance_date = request.args.get('date')
        class_subject_id = request.args.get('class_subject_id', type=int)
        
        if not attendance_date:
            return jsonify({'success': False, 'message': 'date parameter required'}), 400
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        faces = unrecognized_model.get_unrecognized_by_date(
            attendance_date=attendance_date,
            class_subject_id=class_subject_id
        )
        
        return jsonify({'success': True, 'count': len(faces), 'faces': faces}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/<int:face_id>/resolve', methods=['POST'])
def resolve_unrecognized_face(face_id):
    """Resolve an unrecognized face by linking to student"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'success': False, 'message': 'student_id required'}), 400
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        unrecognized_model.resolve_face(
            unrecognized_id=face_id,
            student_id=student_id,
            resolved_by=session.get('user_id')
        )
        
        return jsonify({'success': True, 'message': 'Face resolved successfully'}), 200
    except Exception as e:
        logger.error(f"Error resolving face: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/bulk-resolve', methods=['POST'])
def bulk_resolve_faces():
    """Resolve multiple unrecognized faces at once"""
    try:
        data = request.get_json()
        face_ids = data.get('face_ids', [])
        student_id = data.get('student_id')
        
        if not face_ids or not student_id:
            return jsonify({'success': False, 'message': 'face_ids and student_id required'}), 400
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        count = unrecognized_model.bulk_resolve_faces(
            unrecognized_ids=face_ids,
            student_id=student_id,
            resolved_by=session.get('user_id')
        )
        
        return jsonify({'success': True, 'message': f'Resolved {count} faces', 'count': count}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/<int:face_id>', methods=['DELETE'])
def delete_unrecognized_face(face_id):
    """Delete an unrecognized face"""
    try:
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        unrecognized_model.delete_unrecognized(face_id)
        return jsonify({'success': True, 'message': 'Face deleted successfully'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/unrecognized-faces/cleanup', methods=['POST'])
def cleanup_old_resolved():
    """Cleanup old resolved faces (admin only)"""
    try:
        days = request.args.get('days', 90, type=int)
        
        from models.unrecognized_face import UnrecognizedFace
        unrecognized_model = UnrecognizedFace()
        
        deleted_count = unrecognized_model.cleanup_old_resolved(days=days)
        return jsonify({
            'success': True, 
            'message': f'Cleaned up {deleted_count} old records', 
            'count': deleted_count
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# SYSTEM LOGS
# ============================================================================

@app.route('/api/logs', methods=['GET'])
def get_logs():
    """Get system logs with filters"""
    try:
        log_type = request.args.get('log_type')
        user_type = request.args.get('user_type')
        user_id = request.args.get('user_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = request.args.get('limit', 100, type=int)
        
      
       
        
        logs = log_model.get_logs(
            log_type=log_type,
            user_type=user_type,
            user_id=user_id,
            start_date=start_date,
            end_date=end_date,
            limit=limit
        )
        
        return jsonify({'success': True, 'count': len(logs), 'logs': logs}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/recent', methods=['GET'])
def get_recent_logs():
    """Get most recent logs"""
    try:
        limit = request.args.get('limit', 100, type=int)
        
        from models.log import Log
        log_model = Log()
        
        logs = log_model.get_recent_logs(limit=limit)
        return jsonify({'success': True, 'count': len(logs), 'logs': logs}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/errors', methods=['GET'])
def get_error_logs():
    """Get error logs from last N hours"""
    try:
        hours = request.args.get('hours', 24, type=int)
        limit = request.args.get('limit', 100, type=int)
        
        from models.log import Log
        log_model = Log()
        
        errors = log_model.get_error_logs(hours=hours, limit=limit)
        return jsonify({'success': True, 'count': len(errors), 'errors': errors}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/attendance', methods=['GET'])
def get_attendance_logs():
    """Get attendance-related logs"""
    try:
        date_str = request.args.get('date')
        limit = request.args.get('limit', 100, type=int)
        
        from models.log import Log
        log_model = Log()
        
        if date_str:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            logs = log_model.get_attendance_logs(date=attendance_date, limit=limit)
        else:
            logs = log_model.get_attendance_logs(limit=limit)
        
        return jsonify({'success': True, 'count': len(logs), 'logs': logs}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/admin-actions', methods=['GET'])
def get_admin_actions():
    """Get admin action logs"""
    try:
        limit = request.args.get('limit', 50, type=int)
        
        from models.log import Log
        log_model = Log()
        
        logs = log_model.get_admin_actions(limit=limit)
        return jsonify({'success': True, 'count': len(logs), 'logs': logs}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/by-action', methods=['GET'])
def get_logs_by_action():
    """Get logs filtered by specific action"""
    try:
        action = request.args.get('action')
        limit = request.args.get('limit', 50, type=int)
        
        if not action:
            return jsonify({'success': False, 'message': 'action parameter required'}), 400
        
        from models.log import Log
        log_model = Log()
        
        logs = log_model.get_logs_by_action(action, limit=limit)
        return jsonify({'success': True, 'count': len(logs), 'logs': logs}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/logs/cleanup', methods=['POST'])
def cleanup_old_logs():
    """Delete old logs (admin only)"""
    try:
        days = request.args.get('days', 90, type=int)
        
        from models.log import Log
        log_model = Log()
        
        deleted_count = log_model.cleanup_old_logs(days=days)
        return jsonify({
            'success': True, 
            'message': f'Cleaned up {deleted_count} old logs', 
            'count': deleted_count
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/teacher/attendance/export', methods=['GET'])
def export_attendance_excel():
    """Export attendance to Excel file"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        date = request.args.get('date')
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        # Get attendance records
        if date:
            records = attendance_model.get_attendance_by_date(class_subject_id, date)
        elif start_date and end_date:
            records = attendance_model.get_attendance_by_date_range(
                class_subject_id, start_date, end_date
            )
        else:
            return jsonify({'success': False, 'message': 'Provide date or date range'}), 400
        
        # Get class info for title
        class_info = class_subject_model.get_class_subject_by_id(class_subject_id)
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Title
        ws.merge_cells('A1:G1')
        title = f"{class_info['subject_name']} - {class_info['branch_code']} Year {class_info['year']} Section {class_info['section']}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Date range
        ws.merge_cells('A2:G2')
        if date:
            ws['A2'] = f"Date: {date}"
        else:
            ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['S.No', 'Roll Number', 'Name', 'Date', 'Time', 'Status', 'Confidence']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        from datetime import timedelta
        for idx, record in enumerate(records, start=1):
            row = idx + 4
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=record.get('roll_number', ''))
            ws.cell(row=row, column=3, value=record.get('name', ''))
            
            # Handle date
            att_date = record.get('attendance_date')
            if isinstance(att_date, date):
                ws.cell(row=row, column=4, value=att_date.strftime('%Y-%m-%d'))
            else:
                ws.cell(row=row, column=4, value=str(att_date))
            
            # Handle time (convert timedelta to string)
            att_time = record.get('attendance_time')
            if isinstance(att_time, timedelta):
                total_seconds = int(att_time.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                ws.cell(row=row, column=5, value=f"{hours:02d}:{minutes:02d}:{seconds:02d}")
            else:
                ws.cell(row=row, column=5, value=str(att_time))
            
            ws.cell(row=row, column=6, value=record.get('status', 'present').upper())
            
            # Confidence with color coding
            confidence = record.get('confidence', 0)
            conf_cell = ws.cell(row=row, column=7, value=f"{confidence:.1%}")
            
            if confidence >= 0.7:
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence >= 0.5:
                conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Summary
        summary_row = len(records) + 6
        ws[f'A{summary_row}'] = "Total Records:"
        ws[f'B{summary_row}'] = len(records)
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        # Save to exports folder
        os.makedirs('exports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'attendance_{class_subject_id}_{timestamp}.xlsx'
        filepath = os.path.join('exports', filename)
        
        wb.save(filepath)
        
        # Send file
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    
    
# ============================================================================
# RUN
# ============================================================================

if __name__ == '__main__':
    print("Server starting on http://0.0.0.0:5000")
    app.run(host='0.0.0.0', port=5000, debug=True) 