"""Attendance Routes - Mark attendance and generate reports"""

from flask import request, jsonify, session, send_file
from datetime import datetime, date, timedelta
import logging
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.routes import attendance_bp
from app.services.attendance_service import AttendanceService
from models.attendance import Attendance
from models.class_subject import ClassSubject
from models.student import Student
from models.unrecognized_face import UnrecognizedFace
from services.face_recognition_service import FaceRecognitionService

logger = logging.getLogger(__name__)


def _get_attendance_service():
    """Get attendance service instance"""
    attendance_model = Attendance()
    face_service = FaceRecognitionService()
    student_model = Student()
    unrecognized_model = UnrecognizedFace()
    class_subject_model = ClassSubject()
    
    return AttendanceService(
        attendance_model, 
        face_service, 
        student_model,
        unrecognized_model,
        class_subject_model
    )


@attendance_bp.route('/mark', methods=['POST'])
def mark_attendance():
    """Mark attendance using face recognition"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        service = _get_attendance_service()
        result = service.mark_attendance(
            file=file,
            class_subject_id=class_subject_id,
            attendance_date_str=attendance_date_str,
            session=session
        )
        
        return jsonify({
            'success': True,
            'message': f"Attendance marked. {result['recognized_count']} students present.",
            **result
        }), 200
        
    except Exception as e:
        logger.error(f"Mark attendance error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/mark-batch', methods=['POST'])
def mark_attendance_batch():
    """Mark attendance using multiple images"""
    try:
        if 'files' not in request.files:
            return jsonify({'success': False, 'message': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        if not files or len(files) == 0:
            return jsonify({'success': False, 'message': 'No files selected'}), 400
        
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        skip_quality_check = request.form.get('skip_quality_check', 'false').lower() == 'true'
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        service = _get_attendance_service()
        result = service.mark_attendance_batch(
            files=files,
            class_subject_id=class_subject_id,
            attendance_date_str=attendance_date_str,
            session=session,
            skip_quality_check=skip_quality_check
        )
        
        return jsonify({
            'success': True,
            'message': f"Batch attendance marked. {result['recognized_count']} students present.",
            **result
        }), 200
        
    except Exception as e:
        logger.error(f"Batch attendance error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/mark-v2', methods=['POST'])
def mark_attendance_v2():
    """Enhanced attendance marking with quality check"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        class_subject_id = request.form.get('class_subject_id', type=int)
        attendance_date_str = request.form.get('attendance_date', date.today().isoformat())
        force_process = request.form.get('force_process', 'false').lower() == 'true'
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        service = _get_attendance_service()
        result = service.mark_attendance(
            file=file,
            class_subject_id=class_subject_id,
            attendance_date_str=attendance_date_str,
            session=session,
            force_process=force_process
        )
        
        return jsonify({'success': True, **result}), 200
        
    except Exception as e:
        logger.error(f"Mark attendance v2 error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/manual', methods=['POST'])
def mark_manual():
    """Manually mark attendance"""
    try:
        data = request.get_json()
        
        service = _get_attendance_service()
        attendance_id = service.mark_attendance_manual(
            student_id=data['student_id'],
            class_subject_id=data['class_subject_id'],
            attendance_date_str=data.get('attendance_date'),
            status=data.get('status', 'present'),
            session=session,
            remarks=data.get('remarks', 'Manually marked')
        )
        
        return jsonify({'success': True, 'attendance_id': attendance_id}), 201
        
    except Exception as e:
        logger.error(f"Manual attendance error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/check-quality', methods=['POST'])
def check_image_quality():
    """Check image quality before processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        
        service = _get_attendance_service()
        quality_result = service.check_image_quality(file)
        
        return jsonify({
            'success': True,
            'quality_check': quality_result
        }), 200
        
    except Exception as e:
        logger.error(f"Quality check error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/report', methods=['GET'])
def get_report():
    """Get attendance report"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        service = _get_attendance_service()
        report = service.get_attendance_report_aggregated(class_subject_id, start_date, end_date)
        
        return jsonify({'success': True, 'report': report}), 200
        
    except Exception as e:
        logger.error(f"Report error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/by-date', methods=['GET'])
def get_by_date():
    """Get attendance by date"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        attendance_date = request.args.get('date', date.today().isoformat())
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        service = _get_attendance_service()
        records = service.get_attendance_by_date(class_subject_id, attendance_date)
        
        return jsonify({'success': True, 'records': records}), 200
        
    except Exception as e:
        logger.error(f"Get by date error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@attendance_bp.route('/export', methods=['GET'])
def export_attendance_excel():
    """Export attendance to Excel"""
    try:
        class_subject_id = request.args.get('class_subject_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        date_param = request.args.get('date')
        
        if not class_subject_id:
            return jsonify({'success': False, 'message': 'class_subject_id required'}), 400
        
        attendance_model = Attendance()
        class_subject_model = ClassSubject()
        
        # Get records
        if date_param:
            records = attendance_model.get_attendance_by_date(class_subject_id, date_param)
        elif start_date and end_date:
            records = attendance_model.get_attendance_by_date_range(class_subject_id, start_date, end_date)
        else:
            return jsonify({'success': False, 'message': 'Provide date or date range'}), 400
        
        # Get class info
        class_info = class_subject_model.get_class_subject_by_id(class_subject_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Title
        ws.merge_cells('A1:G1')
        title = f"{class_info['subject_name']} - {class_info['branch_code']} Year {class_info['year']} Section {class_info['section']}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Date range
        ws.merge_cells('A2:G2')
        if date_param:
            ws['A2'] = f"Date: {date_param}"
        else:
            ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['S.No', 'Roll Number', 'Name', 'Date', 'Time', 'Status', 'Confidence']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for idx, record in enumerate(records, start=1):
            row = idx + 4
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=record.get('roll_number', ''))
            ws.cell(row=row, column=3, value=record.get('name', ''))
            
            att_date = record.get('attendance_date')
            if isinstance(att_date, date):
                ws.cell(row=row, column=4, value=att_date.strftime('%Y-%m-%d'))
            else:
                ws.cell(row=row, column=4, value=str(att_date))
            
            att_time = record.get('attendance_time')
            if isinstance(att_time, timedelta):
                total_seconds = int(att_time.total_seconds())
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                ws.cell(row=row, column=5, value=f"{hours:02d}:{minutes:02d}:{seconds:02d}")
            else:
                ws.cell(row=row, column=5, value=str(att_time))
            
            ws.cell(row=row, column=6, value=record.get('status', 'present').upper())
            
            confidence = record.get('confidence', 0)
            conf_cell = ws.cell(row=row, column=7, value=f"{confidence:.1%}")
            
            if confidence >= 0.7:
                conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confidence >= 0.5:
                conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        
        # Save
        os.makedirs('exports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'attendance_{class_subject_id}_{timestamp}.xlsx'
        filepath = os.path.join('exports', filename)
        
        wb.save(filepath)
        
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
