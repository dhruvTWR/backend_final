import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from config import Config

class ExcelExporter:
    def __init__(self, database):
        self.db = database
        self.export_path = Config.EXCEL_EXPORT_PATH
        os.makedirs(self.export_path, exist_ok=True)
    
    def export_attendance(self, date=None, start_date=None, end_date=None):
        """Export attendance records to styled Excel file"""
        try:
            # Get attendance records
            records = self.db.get_attendance(
                date=date,
                start_date=start_date,
                end_date=end_date
            )
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Set column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            
            # Styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = "ATTENDANCE REPORT"
            title_cell.font = Font(bold=True, size=16, color="1F4E78")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30
            
            # Date range info
            ws.merge_cells('A2:F2')
            info_cell = ws['A2']
            if date:
                info_cell.value = f"Date: {date}"
            elif start_date and end_date:
                info_cell.value = f"Period: {start_date} to {end_date}"
            else:
                info_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            info_cell.alignment = Alignment(horizontal="center")
            info_cell.font = Font(italic=True)
            ws.row_dimensions[2].height = 20
            
            # Headers
            headers = ['S.No', 'Name', 'Roll Number', 'Date', 'Time', 'Confidence']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border_style
            
            ws.row_dimensions[4].height = 25
            
            # Data rows
            for idx, record in enumerate(records, start=1):
                row = idx + 4
                
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=record.get('name', 'N/A'))
                ws.cell(row=row, column=3, value=record.get('roll_number', 'N/A'))
                ws.cell(row=row, column=4, value=record.get('attendance_date', 'N/A'))
                ws.cell(row=row, column=5, value=record.get('attendance_time', 'N/A'))
                
                confidence = record.get('confidence', 0)
                confidence_cell = ws.cell(row=row, column=6, value=f"{confidence:.2%}")
                
                # Color code confidence
                if confidence >= 0.8:
                    confidence_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence >= 0.6:
                    confidence_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    confidence_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Apply borders to all cells
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = border_style
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            
            # Summary section
            summary_row = len(records) + 6
            ws.merge_cells(f'A{summary_row}:B{summary_row}')
            summary_cell = ws[f'A{summary_row}']
            summary_cell.value = "Summary"
            summary_cell.font = Font(bold=True, size=14)
            
            ws[f'A{summary_row+1}'] = "Total Records:"
            ws[f'B{summary_row+1}'] = len(records)
            ws[f'A{summary_row+1}'].font = Font(bold=True)
            
            # Get unique students count
            unique_students = len(set(r.get('student_id') for r in records))
            ws[f'A{summary_row+2}'] = "Unique Students:"
            ws[f'B{summary_row+2}'] = unique_students
            ws[f'A{summary_row+2}'].font = Font(bold=True)
            
            # Save file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'attendance_report_{timestamp}.xlsx'
            filepath = os.path.join(self.export_path, filename)
            wb.save(filepath)
            
            print(f"Excel report exported: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Excel export error: {e}")
            raise
    
    def export_student_list(self):
        """Export all students to Excel"""
        try:
            students = self.db.get_all_students()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Students List"
            
            # Headers
            headers = ['S.No', 'Name', 'Roll Number', 'Email', 'Registered Date']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Data
            for idx, student in enumerate(students, start=1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=student['name'])
                ws.cell(row=idx+1, column=3, value=student['roll_number'])
                ws.cell(row=idx+1, column=4, value=student.get('email', 'N/A'))
                ws.cell(row=idx+1, column=5, value=str(student.get('created_at', 'N/A')))
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'students_list_{timestamp}.xlsx'
            filepath = os.path.join(self.export_path, filename)
            wb.save(filepath)
            
            print(f"Student list exported: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Student list export error: {e}")
            raise
    
    def export_daily_summary(self, date=None):
        """Export daily attendance summary"""
        try:
            if not date:
                date = datetime.now().date()
            
            # Get attendance for the date
            records = self.db.get_attendance(date=str(date))
            all_students = self.db.get_all_students()
            
            # Create present/absent lists
            present_ids = {r['student_id'] for r in records}
            present_students = [s for s in all_students if s['id'] in present_ids]
            absent_students = [s for s in all_students if s['id'] not in present_ids]
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Summary {date}"
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = f"Daily Attendance Summary - {date}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Statistics
            ws['A3'] = "Total Students:"
            ws['B3'] = len(all_students)
            ws['A4'] = "Present:"
            ws['B4'] = len(present_students)
            ws['A5'] = "Absent:"
            ws['B5'] = len(absent_students)
            ws['A6'] = "Attendance %:"
            ws['B6'] = f"{(len(present_students)/len(all_students)*100):.1f}%" if all_students else "0%"
            
            for row in range(3, 7):
                ws[f'A{row}'].font = Font(bold=True)
            
            # Present students
            ws['A8'] = "PRESENT STUDENTS"
            ws['A8'].font = Font(bold=True, size=12, color="008000")
            
            ws['A9'] = "S.No"
            ws['B9'] = "Name"
            ws['C9'] = "Roll Number"
            ws['D9'] = "Time"
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}9'].font = Font(bold=True)
                ws[f'{col}9'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            for idx, record in enumerate(records, start=1):
                row = idx + 9
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = record['name']
                ws[f'C{row}'] = record['roll_number']
                ws[f'D{row}'] = record.get('attendance_time', 'N/A')
            
            # Absent students
            absent_start_row = len(records) + 12
            ws[f'A{absent_start_row}'] = "ABSENT STUDENTS"
            ws[f'A{absent_start_row}'].font = Font(bold=True, size=12, color="FF0000")
            
            ws[f'A{absent_start_row+1}'] = "S.No"
            ws[f'B{absent_start_row+1}'] = "Name"
            ws[f'C{absent_start_row+1}'] = "Roll Number"
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{absent_start_row+1}'].font = Font(bold=True)
                ws[f'{col}{absent_start_row+1}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for idx, student in enumerate(absent_students, start=1):
                row = absent_start_row + 1 + idx
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = student['name']
                ws[f'C{row}'] = student['roll_number']
            
            # Adjust columns
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            
            # Save
            filename = f'daily_summary_{date}.xlsx'
            filepath = os.path.join(self.export_path, filename)
            wb.save(filepath)
            
            print(f"Daily summary exported: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Daily summary export error: {e}")
            raise